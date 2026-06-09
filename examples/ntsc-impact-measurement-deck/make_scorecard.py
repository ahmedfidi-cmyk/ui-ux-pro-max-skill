#!/usr/bin/env python3
"""NTSC Participation Impact Scorecard — target vs actual, branded, with formulas."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

EVENING="FF004E43"; GABLE="FF17332F"; HIGHLAND="FF71946A"; ANZAC="FFDEB83B"
SWIRL="FFD5D0C3"; PAPER="FFF6F4EE"; LIGHT="FFEEF2EC"; WHITE="FFFFFFFF"; INK="FF2C3A36"
GREEN_OK="FFE3EFE0"; AMBER="FFFBF3D9"; RED="FFF6E2E2"

thin=Side(style="thin",color="FFCFDAD2")
border=Border(left=thin,right=thin,top=thin,bottom=thin)
F_TITLE=Font(name="Segoe UI",size=15,bold=True,color=EVENING)
F_SUB=Font(name="Segoe UI",size=9,color=HIGHLAND)
F_H=Font(name="Segoe UI",size=10,bold=True,color=WHITE)
F_LBL=Font(name="Segoe UI",size=10,bold=True,color=EVENING)
F=Font(name="Segoe UI",size=10,color=INK)
F_TGT=Font(name="Segoe UI",size=10,bold=True,color=INK)
C=Alignment(horizontal="center",vertical="center",wrap_text=True)
R=Alignment(horizontal="right",vertical="center",wrap_text=True)
fill_head=PatternFill("solid",fgColor=EVENING)
fill_lbl=PatternFill("solid",fgColor=LIGHT)
fill_band=PatternFill("solid",fgColor="FFF4F6F2")
fill_gold=PatternFill("solid",fgColor=AMBER)

wb=openpyxl.Workbook()

# ============================================================
# SHEET 1 — Event Scorecard (single event, detailed)
# ============================================================
ws=wb.active; ws.title="بطاقة التقييم"; ws.sheet_view.rightToLeft=True
ws.sheet_view.showGridLines=False
widths=[3,30,16,16,16,18,22]
for i,w in enumerate(widths,1): ws.column_dimension=None
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

ws.merge_cells("B2:E2"); ws["B2"]="بطاقة قياس أثر المشاركة"; ws["B2"].font=F_TITLE; ws["B2"].alignment=R
ws.merge_cells("B3:E3"); ws["B3"]="Participation Impact Scorecard · NTSC Communications"; ws["B3"].font=F_SUB; ws["B3"].alignment=R
ws.merge_cells("F2:G3"); ws["F2"]="المركز الوطني لسلامة النقل\nNational Transport Safety Center"
ws["F2"].font=Font(name="Segoe UI",size=9,bold=True,color=EVENING); ws["F2"].alignment=C

# event meta
meta=[("اسم الفعالية","Event name"),("مكان وتاريخ المشاركة","Venue & date"),
      ("نوع المشاركة","Participation type"),("مدير المشاركة","Lead")]
r=5
for ar,en in meta:
    ws.merge_cells(f"B{r}:C{r}"); ws[f"B{r}"]=ar; ws[f"B{r}"].font=F_LBL; ws[f"B{r}"].fill=fill_lbl
    ws[f"B{r}"].alignment=R; ws[f"B{r}"].border=border; ws[f"C{r}"].border=border; ws[f"C{r}"].fill=fill_lbl
    ws.merge_cells(f"D{r}:G{r}")
    for col in "DEFG": ws[f"{col}{r}"].border=border
    r+=1
r+=1

# KPI table header
hdr_row=r
heads=["#","المؤشر / KPI","النوع","المستهدف","الناتج الفعلي","نسبة التحقق","التقييم"]
for j,h in enumerate(heads,1):
    c=ws.cell(row=hdr_row,column=j,value=h); c.font=F_H; c.fill=fill_head; c.alignment=C; c.border=border
ws.row_dimensions[hdr_row].height=26

# KPI rows. Target stored as text + numeric helper for ratio.
# columns: A#=1 B=name C=type D=target(text) E=actual(num) F=ratio G=rating
kpis=[("عدد زوّار الجناح (٪ من الزوّار)","نسبي","10",1,"٪"),
      ("مدّة بقاء الزائر (دقائق)","كمّي","5",0,"دقيقة"),
      ("الشراكات المستهدَفة","كمّي","2",0,"شراكة"),
      ("التغطيات الإعلامية الخارجية","كمّي","1",0,"تغطية")]
first=hdr_row+1
for i,(name,typ,tgt,_,unit) in enumerate(kpis):
    rr=first+i
    ws.cell(row=rr,column=1,value=i+1).alignment=C
    ws.cell(row=rr,column=2,value=name).font=F; ws.cell(row=rr,column=2).alignment=R
    ws.cell(row=rr,column=3,value=typ).alignment=C; ws.cell(row=rr,column=3).font=F
    ws.cell(row=rr,column=4,value=int(tgt)).font=F_TGT; ws.cell(row=rr,column=4).alignment=C
    ws.cell(row=rr,column=5).alignment=C   # actual (user fills)
    # ratio formula = actual/target
    ws.cell(row=rr,column=6,value=f"=IF(N(D{rr})=0,\"\",E{rr}/D{rr})").alignment=C
    ws.cell(row=rr,column=6).number_format="0%"
    # rating formula
    ws.cell(row=rr,column=7,value=f'=IF(F{rr}="","—",IF(F{rr}>=1,"محقّق ✔",IF(F{rr}>=0.7,"جزئي ◐","دون المستهدف ✘")))').alignment=C
    ws.cell(row=rr,column=7).font=F
    for j in range(1,8):
        cell=ws.cell(row=rr,column=j); cell.border=border
        if i%2==1 and j not in (5,): cell.fill=fill_band
ws.row_dimensions
last=first+len(kpis)-1

# overall achievement
oar=last+1
ws.merge_cells(f"A{oar}:E{oar}")
ws.cell(row=oar,column=1,value="نسبة التحقق الإجمالية (متوسط) · Overall Achievement").font=F_LBL
ws.cell(row=oar,column=1).alignment=R; ws.cell(row=oar,column=1).fill=fill_gold
ws.merge_cells(f"F{oar}:G{oar}")
ws.cell(row=oar,column=6,value=f"=IFERROR(AVERAGE(F{first}:F{last}),\"\")")
ws.cell(row=oar,column=6).number_format="0%"; ws.cell(row=oar,column=6).font=Font(name="Segoe UI",size=12,bold=True,color=EVENING)
ws.cell(row=oar,column=6).alignment=C; ws.cell(row=oar,column=6).fill=fill_gold
for j in range(1,8): ws.cell(row=oar,column=j).border=border

# data validation note + lessons
ln=oar+2
for title in [("نقاط القوة · Strengths"),("التحديات · Challenges"),("التوصيات وخطة التحسين · Recommendations")]:
    ws.merge_cells(f"B{ln}:G{ln}")
    ws.cell(row=ln,column=2,value=title).font=F_LBL; ws.cell(row=ln,column=2).fill=fill_lbl
    ws.cell(row=ln,column=2).alignment=R;
    for col in "BCDEFG": ws[f"{col}{ln}"].border=border; ws[f"{col}{ln}"].fill=fill_lbl
    ws.merge_cells(f"B{ln+1}:G{ln+1}")
    for col in "BCDEFG": ws[f"{col}{ln+1}"].border=border
    ws.row_dimensions[ln+1].height=30
    ln+=2

# conditional formatting on ratio column F
from openpyxl.formatting.rule import CellIsRule
rng=f"F{first}:F{last}"
ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual",formula=["1"],fill=PatternFill("solid",fgColor=GREEN_OK)))
ws.conditional_formatting.add(rng, CellIsRule(operator="between",formula=["0.7","0.9999"],fill=PatternFill("solid",fgColor=AMBER)))
ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan",formula=["0.7"],fill=PatternFill("solid",fgColor=RED)))

# ============================================================
# SHEET 2 — Events Log (multi-event tracker)
# ============================================================
ws2=wb.create_sheet("سجل الفعاليات"); ws2.sheet_view.rightToLeft=True; ws2.sheet_view.showGridLines=False
ws2.merge_cells("B2:H2"); ws2["B2"]="سجل قياس أثر المشاركات — تتبّع سنوي"; ws2["B2"].font=F_TITLE; ws2["B2"].alignment=R
ws2.merge_cells("B3:H3"); ws2["B3"]="Annual Participation Tracker · target vs. actual"; ws2["B3"].font=F_SUB; ws2["B3"].alignment=R
cols=[("#",4),("الفعالية\nEvent",26),("التاريخ\nDate",13),
      ("الزوّار %\nVisitors (t:10)",13),("المدّة دقائق\nDwell (t:5)",13),
      ("الشراكات\nPartners (t:2)",12),("الإعلام\nMedia (t:1)",11),("التحقق الإجمالي\nAchievement",15)]
hr=5
for j,(h,w) in enumerate(cols,2):
    c=ws2.cell(row=hr,column=j,value=h); c.font=F_H; c.fill=fill_head; c.alignment=C; c.border=border
    ws2.column_dimensions[get_column_letter(j)].width=w
ws2.row_dimensions[hr].height=30
ws2.column_dimensions["A"].width=2
TGT=[10,5,2,1]
for i in range(1,13):  # 12 rows ready
    rr=hr+i
    ws2.cell(row=rr,column=2,value=i).alignment=C
    for j in range(3,9): ws2.cell(row=rr,column=j).alignment=C
    # achievement = average of (actual/target) across the 4 metrics, cols E..H = 5..8
    f=(f'=IFERROR(AVERAGE(E{rr}/10,F{rr}/5,G{rr}/2,H{rr}/1),"")')
    ws2.cell(row=rr,column=9,value=f); ws2.cell(row=rr,column=9).number_format="0%"
    ws2.cell(row=rr,column=9).alignment=C
    for j in range(2,10):
        cell=ws2.cell(row=rr,column=j); cell.border=border; cell.font=F
        if i%2==1: cell.fill=fill_band
# averages row
ar=hr+13
ws2.cell(row=ar,column=2,value="المتوسط · Average").font=F_LBL; ws2.cell(row=ar,column=2).alignment=R; ws2.cell(row=ar,column=2).fill=fill_gold
for j,col in zip(range(5,10),"EFGHI"):
    ws2.cell(row=ar,column=j,value=f"=IFERROR(AVERAGE({col}{hr+1}:{col}{hr+12}),\"\")")
    ws2.cell(row=ar,column=j).alignment=C; ws2.cell(row=ar,column=j).fill=fill_gold
    ws2.cell(row=ar,column=j).font=Font(name="Segoe UI",size=10,bold=True,color=EVENING)
ws2.cell(row=ar,column=9).number_format="0%"
for j in range(2,10): ws2.cell(row=ar,column=j).border=border
ws2.conditional_formatting.add(f"I{hr+1}:I{hr+12}",
    CellIsRule(operator="greaterThanOrEqual",formula=["1"],fill=PatternFill("solid",fgColor=GREEN_OK)))
ws2.conditional_formatting.add(f"I{hr+1}:I{hr+12}",
    CellIsRule(operator="between",formula=["0.7","0.9999"],fill=PatternFill("solid",fgColor=AMBER)))
ws2.conditional_formatting.add(f"I{hr+1}:I{hr+12}",
    CellIsRule(operator="lessThan",formula=["0.7"],fill=PatternFill("solid",fgColor=RED)))

# ============================================================
# SHEET 3 — KPI reference
# ============================================================
ws3=wb.create_sheet("المؤشرات المعتمدة"); ws3.sheet_view.rightToLeft=True; ws3.sheet_view.showGridLines=False
ws3.merge_cells("B2:E2"); ws3["B2"]="المؤشرات والمستهدفات المعتمدة"; ws3["B2"].font=F_TITLE; ws3["B2"].alignment=R
for j,w in zip(range(2,6),[30,14,18,30]): ws3.column_dimensions[get_column_letter(j)].width=w
ws3.column_dimensions["A"].width=2
hh=["المؤشر","النوع","المستهدف","ملاحظات"]
for j,h in enumerate(hh,2):
    c=ws3.cell(row=4,column=j,value=h); c.font=F_H; c.fill=fill_head; c.alignment=C; c.border=border
ref=[("عدد زوّار الجناح","نسبي","٥٪ (٥١ جناحًا أو أكثر) / ١٠٪ (٥٠ أو أقل)","يُحتسب كنسبة من إجمالي زوّار المعرض"),
     ("مدّة بقاء الزائر","كمّي","٢–٥ دقائق","متوسط الزمن داخل الجناح"),
     ("الشراكات المستهدَفة","كمّي","٢","شراكات نوعية"),
     ("التغطيات الإعلامية الخارجية","كمّي","١","من جهة خارج المركز")]
for i,(n,t,tg,note) in enumerate(ref):
    rr=5+i
    vals=[n,t,tg,note]
    for j,v in enumerate(vals,2):
        c=ws3.cell(row=rr,column=j,value=v); c.border=border; c.font=F
        c.alignment=R if j in (2,5) else C
        if i%2==1: c.fill=fill_band
ws3.cell(row=10,column=2,value="ملاحظة: تُجمع البيانات يدويًا عن طريق فريق ممثّلي المركز خلال المشاركة.").font=F_SUB
ws3.cell(row=10,column=2).alignment=R; ws3.merge_cells("B10:E10")

# freeze + protect-ish
ws.sheet_view.zoomScale=110; ws2.sheet_view.zoomScale=110
ws2.freeze_panes="B6"

wb.save("NTSC-Impact-Scorecard.xlsx")
print("saved NTSC-Impact-Scorecard.xlsx with sheets:", wb.sheetnames)
